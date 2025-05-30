import pandas as pd
import re
import io
import json
from datetime import datetime
import streamlit as st

# Try to import required libraries for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def extract_all_tables_from_document(ocr_text, document_name, client=None, model="mistral-large-latest"):
    """
    Extract all possible tables from any document type with enhanced accuracy
    """
    tables = []
    
    # Method 1: Enhanced pattern-based extraction
    pattern_tables = extract_tables_with_enhanced_patterns(ocr_text, document_name)
    tables.extend(pattern_tables)
    
    # Method 2: LLM-enhanced extraction for better accuracy
    if client:
        try:
            llm_tables = extract_tables_with_improved_llm(ocr_text, document_name, client, model)
            # Merge LLM results with pattern results
            tables.extend(llm_tables)
        except Exception as e:
            print(f"LLM extraction failed: {e}")
    
    # Method 3: Specific document type extraction
    specific_tables = extract_specific_document_tables(ocr_text, document_name)
    tables.extend(specific_tables)
    
    return tables

def extract_tables_with_enhanced_patterns(ocr_text, document_name):
    """
    Enhanced pattern-based table extraction with better column detection
    """
    tables = []
    lines = ocr_text.split('\n')
    
    # Look for table-like structures
    potential_table_data = []
    current_table = []
    headers = []
    
    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            if current_table and len(current_table) >= 2:
                # Process accumulated table
                table_df = create_table_from_data(current_table, headers, document_name)
                if table_df is not None:
                    tables.append(table_df)
            current_table = []
            headers = []
            continue
        
        # Split by multiple delimiters
        parts = re.split(r'\s{2,}|\t|\|', line)
        parts = [clean_cell_value(part) for part in parts if part.strip()]
        
        if len(parts) >= 2:
            # Check if this could be headers
            if not headers and is_likely_header_row(parts):
                headers = parts
            else:
                current_table.append(parts)
    
    # Process final table if exists
    if current_table and len(current_table) >= 2:
        table_df = create_table_from_data(current_table, headers, document_name)
        if table_df is not None:
            tables.append(table_df)
    
    return tables

def clean_cell_value(value):
    """Clean cell values by removing unwanted characters"""
    if not value:
        return ""
    
    # Remove dollar signs and other unwanted symbols
    cleaned = str(value).strip()
    cleaned = re.sub(r'^\$+', '', cleaned)  # Remove leading dollar signs
    cleaned = re.sub(r'\$+$', '', cleaned)  # Remove trailing dollar signs
    cleaned = re.sub(r'^\:+', '', cleaned)  # Remove leading colons
    cleaned = re.sub(r'\:+$', '', cleaned)  # Remove trailing colons
    
    return cleaned.strip()

def is_likely_header_row(parts):
    """Determine if a row is likely to be headers"""
    if not parts:
        return False
    
    # Check for common header patterns
    header_indicators = [
        'heavy metal', 'symbol', 'specification', 'vendor', 'ingredient',
        'energy', 'protein', 'fat', 'nutrient', 'value', 'unit',
        'test', 'parameter', 'method', 'result', 'limit',
        'name', 'description', 'code', 'date'
    ]
    
    text_content = ' '.join(parts).lower()
    return any(indicator in text_content for indicator in header_indicators)

def create_table_from_data(table_data, headers, document_name):
    """Create a properly formatted table from extracted data"""
    if not table_data:
        return None
    
    try:
        # Determine consistent column count
        col_counts = [len(row) for row in table_data]
        if not col_counts:
            return None
        
        most_common_cols = max(set(col_counts), key=col_counts.count)
        
        # Filter and pad rows to consistent length
        consistent_rows = []
        for row in table_data:
            if len(row) >= 2:  # At least 2 columns
                padded_row = row[:most_common_cols] + [''] * (most_common_cols - len(row))
                consistent_rows.append(padded_row)
        
        if len(consistent_rows) < 2:
            return None
        
        # Use provided headers or detect from data
        if headers and len(headers) == most_common_cols:
            column_names = [clean_cell_value(h) for h in headers]
        else:
            # Try to detect headers from first row or create generic ones
            first_row = consistent_rows[0]
            if all(not any(char.isdigit() for char in str(cell)) for cell in first_row):
                column_names = [clean_cell_value(cell) for cell in first_row]
                consistent_rows = consistent_rows[1:]  # Remove header row from data
            else:
                column_names = [f'Column_{i+1}' for i in range(most_common_cols)]
        
        # Ensure unique column names
        unique_columns = []
        for i, col in enumerate(column_names):
            if not col or col in unique_columns:
                col = f'Column_{i+1}'
            unique_columns.append(col)
        
        # Create DataFrame
        df = pd.DataFrame(consistent_rows, columns=unique_columns)
        
        # Clean the data
        for col in df.columns:
            df[col] = df[col].apply(clean_cell_value)
        
        # Remove empty rows
        df = df.dropna(how='all')
        df = df[df.iloc[:, 0].astype(str).str.strip() != '']
        
        if df.empty:
            return None
        
        # Determine table type and name
        table_name = determine_table_name(df, document_name)
        table_type = determine_table_type(df)
        
        return {
            'table_name': table_name,
            'dataframe': df,
            'description': f'{table_name} extracted from {document_name}',
            'extraction_method': 'enhanced_pattern'
        }
        
    except Exception as e:
        print(f"Error creating table: {e}")
        return None

def determine_table_name(df, document_name):
    """Determine appropriate table name based on content"""
    # Check column names and content for clues
    columns_text = ' '.join(df.columns).lower()
    content_sample = ' '.join(df.iloc[:3].astype(str).values.flatten()).lower()
    
    if any(word in columns_text or word in content_sample for word in ['heavy', 'metal', 'arsenic', 'lead', 'mercury']):
        return 'Heavy Metals Analysis'
    elif any(word in columns_text or word in content_sample for word in ['energy', 'protein', 'fat', 'carbohydrate', 'nutrition']):
        return 'Nutritional Information'
    elif any(word in columns_text or word in content_sample for word in ['test', 'method', 'specification', 'parameter']):
        return 'Test Specifications'
    elif any(word in columns_text or word in content_sample for word in ['allergen', 'gluten', 'dairy']):
        return 'Allergen Information'
    else:
        return 'Document Data Table'

def determine_table_type(df):
    """Determine table type for categorization"""
    columns_text = ' '.join(df.columns).lower()
    content_sample = ' '.join(df.iloc[:3].astype(str).values.flatten()).lower()
    
    if 'heavy' in columns_text or 'metal' in columns_text:
        return 'heavy_metals'
    elif 'nutrition' in columns_text or 'energy' in columns_text:
        return 'nutritional'
    elif 'test' in columns_text or 'method' in columns_text:
        return 'testing'
    else:
        return 'general'

def extract_tables_with_improved_llm(ocr_text, document_name, client, model):
    """
    Improved LLM-based table extraction with better prompting
    """
    tables = []
    
    try:
        prompt = f"""
Extract ALL tabular data from this document with exact values and proper column names.

Document: {document_name}

Content:
{ocr_text}

Instructions:
1. Find ALL tables in the document
2. Extract EXACT values without adding symbols like $ or :
3. Use the ACTUAL column headers from the document
4. Preserve original formatting of numbers and units
5. Include ALL rows of data

For each table, format as:
TABLE_START
Title: [Descriptive table title]
Headers: [Actual Header 1]|[Actual Header 2]|[Actual Header 3]
[Exact Value 1]|[Exact Value 2]|[Exact Value 3]
[Exact Value 1]|[Exact Value 2]|[Exact Value 3]
TABLE_END

Extract every table you can find with exact values.
"""

        response = client.chat.complete(
            model=model,
            messages=[
                {"role": "system", "content": "You are a precise data extraction specialist. Extract exact values without modification."},
                {"role": "user", "content": prompt}
            ]
        )
        
        llm_response = response.choices[0].message.content
        tables = parse_improved_llm_response(llm_response, document_name)
        
    except Exception as e:
        print(f"LLM table extraction failed: {e}")
    
    return tables

def parse_improved_llm_response(llm_response, document_name):
    """Parse improved LLM response with better accuracy"""
    tables = []
    
    table_blocks = re.findall(r'TABLE_START(.*?)TABLE_END', llm_response, re.DOTALL)
    
    for block in table_blocks:
        try:
            lines = [line.strip() for line in block.strip().split('\n') if line.strip()]
            
            title = "Extracted Table"
            headers = []
            data_rows = []
            
            for line in lines:
                if line.startswith('Title:'):
                    title = line.replace('Title:', '').strip()
                elif line.startswith('Headers:'):
                    headers_line = line.replace('Headers:', '').strip()
                    headers = [clean_cell_value(h) for h in headers_line.split('|')]
                elif '|' in line and not line.startswith(('Title:', 'Headers:')):
                    row_data = [clean_cell_value(cell) for cell in line.split('|')]
                    if len(row_data) == len(headers) and headers:
                        data_rows.append(row_data)
            
            if headers and data_rows:
                df = pd.DataFrame(data_rows, columns=headers)
                tables.append({
                    'table_name': title,
                    'dataframe': df,
                    'description': f'{title} extracted from {document_name}',
                    'extraction_method': 'llm_improved'
                })
                
        except Exception as e:
            print(f"Error parsing LLM table block: {e}")
            continue
    
    return tables

def extract_specific_document_tables(ocr_text, document_name):
    """Extract tables based on specific document patterns"""
    tables = []
    
    # Heavy metals pattern
    if 'heavy metal' in ocr_text.lower() or 'arsenic' in ocr_text.lower():
        heavy_metals_table = extract_heavy_metals_table(ocr_text, document_name)
        if heavy_metals_table:
            tables.append(heavy_metals_table)
    
    # Nutritional data pattern
    if 'nutritional data' in ocr_text.lower() or 'energy' in ocr_text.lower():
        nutrition_table = extract_nutrition_table(ocr_text, document_name)
        if nutrition_table:
            tables.append(nutrition_table)
    
    return tables

def extract_heavy_metals_table(ocr_text, document_name):
    """Extract heavy metals table with specific patterns"""
    lines = ocr_text.split('\n')
    table_data = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Look for heavy metals entries
        if any(metal in line.lower() for metal in ['arsenic', 'lead', 'cadmium', 'mercury', 'copper']):
            parts = re.split(r'\s{2,}|\t', line)
            if len(parts) >= 3:
                metal = clean_cell_value(parts[0])
                symbol = clean_cell_value(parts[1])
                specification = clean_cell_value(parts[2])
                table_data.append([metal, symbol, specification])
    
    if table_data:
        df = pd.DataFrame(table_data, columns=['Heavy Metal', 'Symbol', 'Vendor Ingredient Specification'])
        return {
            'table_name': 'Heavy Metals Analysis',
            'dataframe': df,
            'description': f'Heavy metals specifications from {document_name}',
            'extraction_method': 'specific_heavy_metals'
        }
    
    return None

def extract_nutrition_table(ocr_text, document_name):
    """Extract nutritional data table with specific patterns"""
    lines = ocr_text.split('\n')
    table_data = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Look for nutritional entries
        if any(nutrient in line.lower() for nutrient in ['energy', 'protein', 'fat', 'carbohydrate', 'vitamin', 'mineral']):
            parts = re.split(r'\s{2,}|\t', line)
            if len(parts) >= 3:
                nutrient = clean_cell_value(parts[0])
                value = clean_cell_value(parts[1])
                unit = clean_cell_value(parts[2])
                table_data.append([nutrient, value, unit])
    
    if table_data:
        df = pd.DataFrame(table_data, columns=['Nutrient', 'Value', 'Unit'])
        return {
            'table_name': 'Nutritional Information',
            'dataframe': df,
            'description': f'Nutritional data from {document_name}',
            'extraction_method': 'specific_nutrition'
        }
    
    return None

def create_comprehensive_document_analysis(ocr_text, document_name, client=None, model="mistral-large-latest"):
    """
    Create comprehensive document analysis with all required sections
    """
    if client:
        return create_enhanced_llm_analysis(ocr_text, document_name, client, model)
    else:
        return create_detailed_basic_analysis(ocr_text, document_name)

def create_enhanced_llm_analysis(ocr_text, document_name, client, model):
    """Create detailed analysis using LLM with comprehensive extraction"""
    try:
        prompt = f"""
Analyze this document comprehensively and extract ALL relevant information.

Document: {document_name}

Content:
{ocr_text}

Provide a detailed analysis in this EXACT format:

Summary:
[Write a comprehensive 3-4 sentence summary describing what this document is about, its purpose, the type of data it contains, and key findings. Be specific about the document type and content.]

Product Information:
[Extract ALL product details including:
- Product names and codes
- Company/manufacturer information
- Dates (creation, testing, etc.)
- Document IDs or reference numbers
- Contact information
- Any other identifying information]

Observations:
[Provide detailed analytical insights including:
- Key data patterns or trends
- Compliance or regulatory information
- Testing methods or procedures mentioned
- Important specifications or limits
- Quality control information
- Any notable findings or conclusions]

Be thorough and extract ALL available information from the document.
"""

        response = client.chat.complete(
            model=model,
            messages=[
                {"role": "system", "content": "You are a comprehensive document analysis expert. Extract ALL relevant information thoroughly."},
                {"role": "user", "content": prompt}
            ]
        )
        
        analysis_text = response.choices[0].message.content
        return parse_comprehensive_analysis(analysis_text, document_name)
        
    except Exception as e:
        print(f"LLM analysis failed: {e}")
        return create_detailed_basic_analysis(ocr_text, document_name)

def parse_comprehensive_analysis(analysis_text, document_name):
    """Parse comprehensive analysis into structured sections"""
    sections = {
        'document_name': document_name,
        'processing_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'summary': '',
        'product_information': '',
        'observations': ''
    }
    
    # Split into sections
    lines = analysis_text.split('\n')
    current_section = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if line.lower().startswith('summary:'):
            current_section = 'summary'
            content = line.replace('Summary:', '').strip()
            if content:
                sections['summary'] = content
        elif line.lower().startswith('product information:'):
            current_section = 'product_information'
            content = line.replace('Product Information:', '').strip()
            if content:
                sections['product_information'] = content
        elif line.lower().startswith('observations:'):
            current_section = 'observations'
            content = line.replace('Observations:', '').strip()
            if content:
                sections['observations'] = content
        elif current_section:
            # Add to current section
            if sections[current_section]:
                sections[current_section] += '\n' + line
            else:
                sections[current_section] = line
    
    return sections

def create_detailed_basic_analysis(ocr_text, document_name):
    """Create detailed basic analysis without LLM"""
    lines = ocr_text.split('\n')
    
    # Extract comprehensive information
    companies = []
    products = []
    dates = []
    codes = []
    technical_data = []
    contact_info = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Company detection
        if any(keyword in line.lower() for keyword in ['inc.', 'corp', 'ltd', 'pt.', 'bv', 'gmbh', 'international', 'flavors', 'fragrances']):
            companies.append(line)
            
        # Date detection
        date_match = re.search(r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b|\b\w+\s+\d{1,2},?\s+\d{4}\b', line)
        if date_match:
            dates.append(line)
            
        # Product code detection
        if re.search(r'\bSC\d+\b|\b\d{8,}\b', line):
            codes.append(line)
            
        # Technical data detection
        if any(char.isdigit() for char in line) and any(unit in line.lower() for unit in ['ppm', 'mg', 'kg', '%', 'cal', 'kj', 'g/']):
            technical_data.append(line)
        
        # Contact information
        if any(keyword in line.lower() for keyword in ['phone', 'email', 'address', 'jakarta', 'indonesia']):
            contact_info.append(line)
    
    # Build comprehensive analysis
    summary = f"The document '{document_name}' is a "
    
    if 'heavy metal' in ocr_text.lower():
        summary += "heavy metals compliance document that provides vendor ingredient specifications for heavy metal limits. "
    elif 'nutritional' in ocr_text.lower():
        summary += "nutritional data document providing calculated nutritional information per 100 grams. "
    else:
        summary += "technical document containing structured data and specifications. "
    
    if companies:
        summary += f"The document is from {companies[0]} and "
    
    if technical_data:
        summary += f"contains {len(technical_data)} technical specifications with detailed measurements and limits."
    
    # Product information
    product_info = ""
    if companies:
        product_info += f"Company: {companies[0]}\n"
    if dates:
        product_info += f"Document Date: {dates[0]}\n"
    if codes:
        for code in codes[:3]:
            product_info += f"Product Code: {code}\n"
    if contact_info:
        for contact in contact_info[:2]:
            product_info += f"Contact: {contact}\n"
    
    # Observations
    observations = ""
    if 'heavy metal' in ocr_text.lower():
        observations += "This document establishes heavy metal specifications for vendor ingredients, including limits for arsenic, lead, cadmium, mercury, and copper. "
    if 'nutritional' in ocr_text.lower():
        observations += "The nutritional data has been calculated from supplier raw material data and nutritional tables. "
    
    observations += f"The document contains {len(technical_data)} technical data entries with specific measurements and compliance requirements. "
    
    if 'random checks' in ocr_text.lower():
        observations += "The document mentions random testing procedures for quality control. "
    
    return {
        'document_name': document_name,
        'processing_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'summary': summary,
        'product_information': product_info,
        'observations': observations
    }

def create_professional_excel_export(ocr_results, file_names, client=None, model="mistral-large-latest"):
    """
    Create professional Excel export with comprehensive data extraction
    """
    if not EXCEL_AVAILABLE:
        return None
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        all_tables_for_consolidation = []
        all_analyses_for_consolidation = []
        
        # Process each document
        for idx, (ocr_text, file_name) in enumerate(zip(ocr_results, file_names)):
            # Clean sheet name for Excel
            sheet_name = re.sub(r'[^\w\s-]', '', file_name)[:31]
            if not sheet_name:
                sheet_name = f"Document_{idx+1}"
            
            # Extract tables and analysis
            tables = extract_all_tables_from_document(ocr_text, file_name, client, model)
            analysis = create_comprehensive_document_analysis(ocr_text, file_name, client, model)
            
            # Create individual document sheet
            create_professional_document_sheet(writer, sheet_name, file_name, tables, analysis)
            
            # Store for consolidated sheet
            all_tables_for_consolidation.extend(tables)
            all_analyses_for_consolidation.append(analysis)
        
        # Create consolidated sheet
        create_consolidated_summary_sheet(writer, all_tables_for_consolidation, all_analyses_for_consolidation)
    
    output.seek(0)
    return output.getvalue()

def create_professional_document_sheet(writer, sheet_name, file_name, tables, analysis):
    """
    Create individual document sheet with exact format and comprehensive content
    """
    worksheet = writer.book.create_sheet(title=sheet_name)
    
    # Set column widths for optimal display
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 2  # Spacer
    worksheet.column_dimensions['E'].width = 60
    
    current_row = 1
    
    # Add tables on the left side (columns A-C)
    if tables:
        for table_idx, table_info in enumerate(tables):
            df = table_info['dataframe']
            
            # Add table headers with professional formatting
            for col_idx, column in enumerate(df.columns):
                cell = worksheet.cell(row=current_row, column=col_idx + 1, value=column)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Add table data with proper formatting
            for _, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    cell = worksheet.cell(row=current_row, column=col_idx + 1, value=str(value))
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    # Alternate row colors for better readability
                    if current_row % 2 == 0:
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                current_row += 1
            
            current_row += 2  # Space between tables
    
    # Add comprehensive summary on the right side (starting column E)
    summary_start_row = 2
    
    # Summary section with content
    summary_cell = worksheet.cell(row=summary_start_row, column=5, value="Summary:")
    summary_cell.font = Font(bold=True, size=14, color="FFFFFF")
    summary_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Summary content with proper text wrapping
    summary_content = analysis.get('summary', 'No summary available')
    summary_text_cell = worksheet.cell(row=summary_start_row + 1, column=5, value=summary_content)
    summary_text_cell.alignment = Alignment(wrap_text=True, vertical='top')
    summary_text_cell.font = Font(size=11)
    
    # Product Information section
    product_start_row = summary_start_row + 10
    product_header_cell = worksheet.cell(row=product_start_row, column=5, value="Product Information:")
    product_header_cell.font = Font(bold=True, size=14, color="FFFFFF")
    product_header_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    product_header_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    product_content = analysis.get('product_information', 'No product information available')
    product_text_cell = worksheet.cell(row=product_start_row + 1, column=5, value=product_content)
    product_text_cell.alignment = Alignment(wrap_text=True, vertical='top')
    product_text_cell.font = Font(size=11)
    
    # Observations section
    obs_start_row = product_start_row + 15
    obs_header_cell = worksheet.cell(row=obs_start_row, column=5, value="Observations:")
    obs_header_cell.font = Font(bold=True, size=14, color="FFFFFF")
    obs_header_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    obs_header_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    obs_content = analysis.get('observations', 'No observations available')
    obs_text_cell = worksheet.cell(row=obs_start_row + 1, column=5, value=obs_content)
    obs_text_cell.alignment = Alignment(wrap_text=True, vertical='top')
    obs_text_cell.font = Font(size=11)
    
    # Add professional border around summary section
    for row in range(summary_start_row, obs_start_row + 20):
        cell = worksheet.cell(row=row, column=5)
        if cell.value is None:
            cell.value = ""
        cell.border = Border(
            left=Side(style='medium', color='4472C4'),
            right=Side(style='medium', color='4472C4'),
            top=Side(style='medium', color='4472C4') if row == summary_start_row else Side(style='thin'),
            bottom=Side(style='medium', color='4472C4') if row == obs_start_row + 19 else Side(style='thin')
        )

def create_consolidated_summary_sheet(writer, all_tables, all_analyses):
    """
    Create consolidated sheet with all data organized professionally
    """
    worksheet = writer.book.create_sheet(title="Consolidated_Summary")
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 60
    
    current_row = 1
    
    # Header
    header_cell = worksheet.cell(row=current_row, column=1, value="CONSOLIDATED DATA SUMMARY - ALL DOCUMENTS")
    header_cell.font = Font(bold=True, size=16, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 3
    
    # All tables section
    if all_tables:
        tables_header = worksheet.cell(row=current_row, column=1, value="ALL EXTRACTED TABLES")
        tables_header.font = Font(bold=True, size=14, color="FFFFFF")
        tables_header.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2
        
        for table in all_tables:
            # Table title
            title_cell = worksheet.cell(row=current_row, column=1, value=f"{table['table_name']}")
            title_cell.font = Font(bold=True, size=12)
            current_row += 1
            
            # Table data
            df = table['dataframe']
            
            # Headers
            for col_idx, column in enumerate(df.columns):
                cell = worksheet.cell(row=current_row, column=col_idx + 1, value=column)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1
            
            # Data rows
            for _, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    cell = worksheet.cell(row=current_row, column=col_idx + 1, value=str(value))
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                current_row += 1
            
            current_row += 2
    
    # All analyses section
    current_row += 3
    analyses_header = worksheet.cell(row=current_row, column=1, value="ALL DOCUMENT ANALYSES")
    analyses_header.font = Font(bold=True, size=14, color="FFFFFF")
    analyses_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2
    
    for analysis in all_analyses:
        doc_header = worksheet.cell(row=current_row, column=1, value=f"Document: {analysis['document_name']}")
        doc_header.font = Font(bold=True, size=12)
        current_row += 1
        
        # Summary
        worksheet.cell(row=current_row, column=1, value="Summary:")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=analysis.get('summary', ''))
        worksheet.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        current_row += 1
        
        # Product Info
        worksheet.cell(row=current_row, column=1, value="Product Info:")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=analysis.get('product_information', ''))
        worksheet.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        current_row += 1
        
        # Observations
        worksheet.cell(row=current_row, column=1, value="Observations:")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=analysis.get('observations', ''))
        worksheet.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        current_row += 2
    
    # Remove default sheet
    if 'Sheet' in writer.book.sheetnames:
        writer.book.remove(writer.book['Sheet'])

# Legacy compatibility functions (keeping for backward compatibility)
def extract_tabular_data_from_text_with_llm(ocr_text, document_name, client=None, model="mistral-large-latest"):
    """Legacy function - redirects to new system"""
    return extract_all_tables_from_document(ocr_text, document_name, client, model)

def create_enhanced_document_analysis_with_llm(ocr_text, document_name, client=None, model="mistral-large-latest"):
    """Legacy function - redirects to new system"""
    return create_comprehensive_document_analysis(ocr_text, document_name, client, model)

def create_enhanced_excel_export(ocr_results, file_names, analysis_results=None):
    """Legacy function - redirects to new system"""
    return create_professional_excel_export(ocr_results, file_names)

def generate_tabular_data_with_llm(ocr_results, file_names, client, model="mistral-large-latest"):
    """
    Generate tabular data using LLM with proper progress indication
    """
    # Create progress bar and status text
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Build combined context with proper progress indication
    combined_context = ""
    
    for idx, (ocr_text, file_name) in enumerate(zip(ocr_results, file_names)):
        # Update progress - show current document being processed
        progress = idx / len(ocr_results)  # 0/3, 1/3, 2/3
        progress_bar.progress(progress)
        status_text.text(f"🔄 Processing document {idx+1}/{len(ocr_results)}: {file_name}")
        
        # Add document to context
        combined_context += f"\n\n=== DOCUMENT {idx+1}: {file_name} ===\n\n{ocr_text}\n\n"
        
        # Add a small delay to show progression
        import time
        time.sleep(0.5)
    
    # Update to show LLM processing
    progress_bar.progress(0.9)
    status_text.text("🤖 Analyzing all documents with AI...")
    
    # Create prompt for LLM to extract tabular data
    prompt = f"""
You are a data extraction specialist. Extract ALL tabular data from the provided documents and format them for Excel export.

DOCUMENTS CONTENT:
{combined_context}

INSTRUCTIONS:
1. Extract ALL tables from ALL documents
2. For each table, provide:
   - A descriptive table name
   - Proper column headers (use actual headers from documents)
   - All data rows with clean values (no $ symbols or unwanted characters)
   - Source document name

3. Also provide document analysis for each document:
   - Summary: Brief description of what the document contains
   - Product Information: Company names, product codes, dates, contact info
   - Observations: Key findings, compliance info, testing methods

4. Format your response as JSON with this structure:
{{
    "tables": [
        {{
            "table_name": "Heavy Metals Analysis",
            "source_document": "document_name.pdf",
            "headers": ["Heavy Metal", "Symbol", "Vendor Ingredient Specification"],
            "data": [
                ["Arsenic", "As", "<=3 ppm"],
                ["Lead", "Pb", "<=10 ppm"]
            ]
        }}
    ],
    "document_analyses": [
        {{
            "document_name": "document_name.pdf",
            "summary": "This document provides heavy metal specifications...",
            "product_information": "Company: IFF International\\nDate: December 16, 2019\\nProduct: Heavy Metal Standards",
            "observations": "Document establishes vendor ingredient specifications for heavy metals including arsenic, lead, cadmium, mercury, and copper limits."
        }}
    ]
}}

Extract ALL tables and provide comprehensive analysis. Use clean values without symbols like $ or :.
"""

    try:
        response = client.chat.complete(
            model=model,
            messages=[
                {"role": "system", "content": "You are a precise data extraction specialist. Extract tabular data and format as JSON."},
                {"role": "user", "content": prompt}
            ]
        )
        
        llm_response = response.choices[0].message.content
        result = parse_llm_tabular_response(llm_response)
        
        # Complete progress bar
        progress_bar.progress(1.0)
        status_text.text("✅ Document processing complete!")
        
        # Clean up progress indicators after a short delay
        import time
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
        return result
        
    except Exception as e:
        print(f"LLM tabular extraction failed: {e}")
        
        # Complete progress bar even on error
        progress_bar.progress(1.0)
        status_text.text("❌ Processing failed!")
        
        # Clean up progress indicators
        import time
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
        return create_fallback_response(ocr_results, file_names)

def parse_llm_tabular_response(llm_response):
    """
    Parse LLM response to extract tables and analyses
    """
    try:
        # Try to extract JSON from the response
        json_match = re.search(r'\{.*\}', llm_response, re.DOTALL)
        if json_match:
            json_str = json_match.group()
            data = json.loads(json_str)
        else:
            # If no JSON found, try to parse the response manually
            return parse_manual_response(llm_response)
        
        # Convert to our expected format
        tables = []
        for table_data in data.get('tables', []):
            if table_data.get('headers') and table_data.get('data'):
                df = pd.DataFrame(table_data['data'], columns=table_data['headers'])
                tables.append({
                    'table_name': table_data.get('table_name', 'Extracted Table'),
                    'dataframe': df,
                    'source_document': table_data.get('source_document', 'Unknown'),
                    'description': f"Table extracted from {table_data.get('source_document', 'document')}"
                })
        
        analyses = data.get('document_analyses', [])
        
        return {
            'tables': tables,
            'analyses': analyses,
            'success': True
        }
        
    except Exception as e:
        print(f"Error parsing LLM response: {e}")
        return parse_manual_response(llm_response)

def parse_manual_response(llm_response):
    """
    Manually parse LLM response if JSON parsing fails
    """
    tables = []
    analyses = []
    
    # Try to extract table-like structures from the response
    lines = llm_response.split('\n')
    current_table_data = []
    current_headers = []
    table_name = "Extracted Table"
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Look for table indicators
        if 'table' in line.lower() and ':' in line:
            table_name = line.split(':')[1].strip()
        elif '|' in line and len(line.split('|')) >= 2:
            parts = [part.strip() for part in line.split('|') if part.strip()]
            if not current_headers:
                current_headers = parts
            else:
                current_table_data.append(parts)
        elif current_table_data and current_headers:
            # End of table, process it
            if len(current_table_data) > 0:
                df = pd.DataFrame(current_table_data, columns=current_headers)
                tables.append({
                    'table_name': table_name,
                    'dataframe': df,
                    'source_document': 'Unknown',
                    'description': f"Table extracted from document"
                })
            current_table_data = []
            current_headers = []
            table_name = "Extracted Table"
    
    # Process final table if exists
    if current_table_data and current_headers:
        df = pd.DataFrame(current_table_data, columns=current_headers)
        tables.append({
            'table_name': table_name,
            'dataframe': df,
            'source_document': 'Unknown',
            'description': f"Table extracted from document"
        })
    
    return {
        'tables': tables,
        'analyses': analyses,
        'success': len(tables) > 0
    }

def create_fallback_response(ocr_results, file_names):
    """
    Create fallback response if LLM fails
    """
    tables = []
    analyses = []
    
    for idx, (ocr_text, file_name) in enumerate(zip(ocr_results, file_names)):
        # Simple pattern-based extraction as fallback
        lines = ocr_text.split('\n')
        table_data = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Look for lines with multiple columns
            parts = re.split(r'\s{2,}|\t', line)
            if len(parts) >= 2:
                cleaned_parts = [part.strip() for part in parts if part.strip()]
                if len(cleaned_parts) >= 2:
                    table_data.append(cleaned_parts)
        
        if table_data and len(table_data) > 1:
            # Use first row as headers
            headers = table_data[0]
            data_rows = table_data[1:]
            
            # Ensure consistent column count
            max_cols = len(headers)
            consistent_data = []
            for row in data_rows:
                if len(row) == max_cols:
                    consistent_data.append(row)
            
            if consistent_data:
                df = pd.DataFrame(consistent_data, columns=headers)
                tables.append({
                    'table_name': f'Data from {file_name}',
                    'dataframe': df,
                    'source_document': file_name,
                    'description': f"Table extracted from {file_name}"
                })
        
        # Basic analysis
        analyses.append({
            'document_name': file_name,
            'summary': f"Document contains {len(ocr_text)} characters of text data.",
            'product_information': "Information extracted using fallback method.",
            'observations': "Basic pattern-based extraction was used."
        })
    
    return {
        'tables': tables,
        'analyses': analyses,
        'success': len(tables) > 0
    }

def create_excel_from_extracted_data(extracted_data, file_names):
    """
    Create Excel file from extracted tabular data
    """
    if not EXCEL_AVAILABLE:
        return None
    
    if not extracted_data.get('success') or not extracted_data.get('tables'):
        return None
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        tables = extracted_data['tables']
        analyses = extracted_data['analyses']
        
        # Group tables by source document
        tables_by_document = {}
        for table in tables:
            doc_name = table['source_document']
            if doc_name not in tables_by_document:
                tables_by_document[doc_name] = []
            tables_by_document[doc_name].append(table)
        
        # Create individual sheets for each document
        for doc_name, doc_tables in tables_by_document.items():
            # Clean sheet name
            sheet_name = re.sub(r'[^\w\s-]', '', doc_name)[:31]
            if not sheet_name:
                sheet_name = f"Document_{len(tables_by_document)}"
            
            # Find corresponding analysis
            doc_analysis = None
            for analysis in analyses:
                if analysis['document_name'] == doc_name:
                    doc_analysis = analysis
                    break
            
            if not doc_analysis:
                doc_analysis = {
                    'summary': 'No analysis available',
                    'product_information': 'No product information available',
                    'observations': 'No observations available'
                }
            
            create_document_sheet_simple(writer, sheet_name, doc_tables, doc_analysis)
        
        # Create consolidated sheet
        create_consolidated_sheet_simple(writer, tables, analyses)
    
    output.seek(0)
    return output.getvalue()

def create_document_sheet_simple(writer, sheet_name, tables, analysis):
    """
    Create individual document sheet with simple, clean format
    """
    worksheet = writer.book.create_sheet(title=sheet_name)
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 3  # Spacer
    worksheet.column_dimensions['E'].width = 60
    
    current_row = 1
    
    # Add tables on the left side (columns A-C)
    for table in tables:
        df = table['dataframe']
        
        # Table title
        title_cell = worksheet.cell(row=current_row, column=1, value=table['table_name'])
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        # Headers
        for col_idx, column in enumerate(df.columns):
            cell = worksheet.cell(row=current_row, column=col_idx + 1, value=column)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1
        
        # Data rows
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row):
                cell = worksheet.cell(row=current_row, column=col_idx + 1, value=str(value))
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # Alternate row colors
                if current_row % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            current_row += 1
        
        current_row += 2  # Space between tables
    
    # Add analysis on the right side (column E)
    analysis_start_row = 2
    
    # Summary
    summary_header = worksheet.cell(row=analysis_start_row, column=5, value="Summary:")
    summary_header.font = Font(bold=True, size=14, color="FFFFFF")
    summary_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary_content = worksheet.cell(row=analysis_start_row + 1, column=5, value=analysis['summary'])
    summary_content.alignment = Alignment(wrap_text=True, vertical='top')
    summary_content.font = Font(size=11)
    
    # Product Information
    product_start_row = analysis_start_row + 8
    product_header = worksheet.cell(row=product_start_row, column=5, value="Product Information:")
    product_header.font = Font(bold=True, size=14, color="FFFFFF")
    product_header.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    product_content = worksheet.cell(row=product_start_row + 1, column=5, value=analysis['product_information'])
    product_content.alignment = Alignment(wrap_text=True, vertical='top')
    product_content.font = Font(size=11)
    
    # Observations
    obs_start_row = product_start_row + 8
    obs_header = worksheet.cell(row=obs_start_row, column=5, value="Observations:")
    obs_header.font = Font(bold=True, size=14, color="FFFFFF")
    obs_header.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    obs_content = worksheet.cell(row=obs_start_row + 1, column=5, value=analysis['observations'])
    obs_content.alignment = Alignment(wrap_text=True, vertical='top')
    obs_content.font = Font(size=11)

def create_consolidated_sheet_simple(writer, all_tables, all_analyses):
    """
    Create consolidated sheet with all data
    """
    worksheet = writer.book.create_sheet(title="Consolidated_Summary")
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 50
    
    current_row = 1
    
    # Header
    header_cell = worksheet.cell(row=current_row, column=1, value="CONSOLIDATED DATA SUMMARY")
    header_cell.font = Font(bold=True, size=16, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 3
    
    # All tables section
    if all_tables:
        tables_header = worksheet.cell(row=current_row, column=1, value="ALL EXTRACTED TABLES")
        tables_header.font = Font(bold=True, size=14, color="FFFFFF")
        tables_header.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2
        
        for table in all_tables:
            # Table title
            title_cell = worksheet.cell(row=current_row, column=1, value=f"{table['table_name']} (from {table['source_document']})")
            title_cell.font = Font(bold=True, size=12)
            current_row += 1
            
            df = table['dataframe']
            
            # Headers
            for col_idx, column in enumerate(df.columns):
                cell = worksheet.cell(row=current_row, column=col_idx + 1, value=column)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            current_row += 1
            
            # Data
            for _, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    worksheet.cell(row=current_row, column=col_idx + 1, value=str(value))
                current_row += 1
            
            current_row += 2
    
    # All analyses section
    current_row += 2
    analyses_header = worksheet.cell(row=current_row, column=1, value="DOCUMENT ANALYSES")
    analyses_header.font = Font(bold=True, size=14, color="FFFFFF")
    analyses_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2
    
    for analysis in all_analyses:
        doc_header = worksheet.cell(row=current_row, column=1, value=f"Document: {analysis['document_name']}")
        doc_header.font = Font(bold=True, size=12)
        current_row += 1
        
        # Summary
        worksheet.cell(row=current_row, column=1, value="Summary:")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=analysis.get('summary', ''))
        current_row += 1
        
        # Product Info
        worksheet.cell(row=current_row, column=1, value="Product Info:")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=analysis.get('product_information', ''))
        current_row += 1
        
        # Observations
        worksheet.cell(row=current_row, column=1, value="Observations:")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=analysis.get('observations', ''))
        current_row += 2
    
    # Remove default sheet
    if 'Sheet' in writer.book.sheetnames:
        writer.book.remove(writer.book['Sheet'])

def create_simple_excel_export(ocr_results, file_names, client, model="mistral-large-latest"):
    """
    Main function to create Excel export using simple LLM approach
    """
    # Step 1: Extract tabular data using LLM
    extracted_data = generate_tabular_data_with_llm(ocr_results, file_names, client, model)
    
    # Step 2: Create Excel file from extracted data
    if extracted_data.get('success'):
        with st.spinner("📊 Creating Excel report..."):
            excel_data = create_excel_from_extracted_data(extracted_data, file_names)
        
        if excel_data:
            return excel_data
        else:
            st.error("❌ Failed to create Excel file")
            return None
    else:
        st.error("❌ No tabular data could be extracted from any documents")
        return None
# Legacy compatibility function
def create_professional_excel_export(ocr_results, file_names, client=None, model="mistral-large-latest"):
    """
    Legacy function - redirects to new simple approach
    """
    if client:
        return create_simple_excel_export(ocr_results, file_names, client, model)
    else:
        return None
