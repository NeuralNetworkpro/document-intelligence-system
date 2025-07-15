import pandas as pd
import io
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

def parse_mismatches_for_update(report):
    """
    Parses the 'mismatches' section of a single AI report to extract
    field names and their correct source values.
    """
    updates = {}
    mismatches = report.get("mismatches", [])
    
    for item in mismatches:
        field_name = item.get("field_name")
        source_value = item.get("source_value")
        # We only want to update with valid, non-placeholder values
        if field_name and source_value and source_value not in ["N/A", "Not specified", "Not explicitly mentioned"]:
            updates[field_name] = source_value
            
    return updates


def create_updated_excel(master_file_bytes, comparison_results):
    """
    Loads the original master Excel file, updates it with values from the
    comparison results, highlights changes, adds comments with old values,
    and returns the bytes of the new Excel file.
    """
    try:
        # Load the workbook from the uploaded file's bytes
        master_workbook = openpyxl.load_workbook(io.BytesIO(master_file_bytes))
        # Assume we are working with the first sheet
        sheet = master_workbook.active

        # Define a fill color for highlighting updated cells (e.g., yellow)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # First, gather all updates from all document comparisons.
        all_updates = {}
        for result in comparison_results:
            report_data = result.get("report", {})
            if report_data and "error" not in report_data:
                updates_from_doc = parse_mismatches_for_update(report_data)
                all_updates.update(updates_from_doc)

        if not all_updates:
            # No valid mismatches to update
            return None

        # Now, iterate through the Excel sheet and apply the updates.
        # We assume the attribute names are in Column A (index 1).
        for row in range(1, sheet.max_row + 1):
            cell_A = sheet.cell(row=row, column=1)
            if cell_A.value:
                attribute_name = str(cell_A.value).strip()
                
                # If this attribute was found in our list of updates
                if attribute_name in all_updates:
                    cell_B = sheet.cell(row=row, column=2)
                    original_value = cell_B.value
                    new_value = all_updates[attribute_name]

                    # Update the cell value
                    cell_B.value = new_value
                    
                    # Highlight the cell
                    cell_B.fill = highlight_fill
                    
                    # Add a comment with the original value
                    comment_text = f"Original Value: {original_value}"
                    cell_B.comment = Comment(comment_text, "Data Intelligence System")

        # FIX: Use an in-memory stream (BytesIO) to save the workbook
        virtual_workbook_stream = io.BytesIO()
        master_workbook.save(virtual_workbook_stream)
        # Return the bytes from the stream by seeking to the beginning and reading
        virtual_workbook_stream.seek(0)
        return virtual_workbook_stream.getvalue()

    except Exception as e:
        # Return the error to be displayed in the UI
        return f"Error creating updated Excel file: {e}"

